import json
import numpy as np
import os
import argparse
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def get_metrics_from_log(prompt_type, log_phase=0):
    """Get test pass rate and pass@k directly from log file."""
    log_file_pattern1 = os.path.join(SCRIPT_DIR, 'log', 'record', '%s_model_%s_topn_%s_temperature_%s.0.log_%s' % (experiment, model, topn, temperature, log_phase))
    log_file_pattern2 = os.path.join(SCRIPT_DIR, 'log', 'record', '%s_model_%s_topn_%s_temperature_%s.0.log_0' % (experiment, model, topn, temperature))
    
    log_file = None
    if os.path.exists(log_file_pattern1):
        log_file = log_file_pattern1
    elif os.path.exists(log_file_pattern2):
        log_file = log_file_pattern2
    else:
        raise FileNotFoundError(f'Log file not found. Tried:\n  - {log_file_pattern1}\n  - {log_file_pattern2}')
    
    total_passed_cases = 0
    total_cases = 0
    problem_pass_status = []
    
    with open(log_file, 'r', encoding='utf-8') as f:
        for line in f:
            entry = json.loads(line)
            entry_prompt_type = entry.get('prompt_type', '')
            
            if prompt_type:
                if prompt_type == 'prompt':
                    if entry_prompt_type != 'prompt' and entry_prompt_type != '':
                        continue
                else:
                    if entry_prompt_type != prompt_type:
                        continue
            
            if entry['code_candidates']:
                code_candidate = entry['code_candidates'][0]
                passed_case = code_candidate.get('passed_case', [])
                case_status = code_candidate.get('case_status', [])
                
                num_passed = len(passed_case)
                num_total = len(case_status)
                
                total_passed_cases += num_passed
                total_cases += num_total
                
                problem_passes = (num_total > 0 and num_passed == num_total)
                problem_pass_status.append(1.0 if problem_passes else 0.0)
    
    return total_passed_cases, total_cases, problem_pass_status


def calculate_metrics(prompt_type, log_phase=0):
    """Calculate test pass rate and pass@k metrics."""
    total_passed_cases, total_cases, problem_pass_status = get_metrics_from_log(prompt_type, log_phase)
    
    if total_cases > 0:
        test_pass_rate = total_passed_cases / total_cases
    else:
        test_pass_rate = 0.0
    
    if len(problem_pass_status) > 0:
        passk = np.mean(problem_pass_status)
    else:
        passk = 0.0
    
    return {
        'test_pass_rate': test_pass_rate,
        'pass@k': passk,
        'total_passed_cases': total_passed_cases,
        'total_cases': total_cases,
        'num_problems': len(problem_pass_status),
        'num_passed_problems': sum(problem_pass_status)
    }


def store_results_xlsx(metrics, file_suffix, suppress_warnings=False):
    """Store results in Excel file."""
    import warnings
    if suppress_warnings:
        warnings.filterwarnings('ignore')
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1).value = "Pass@k (%)"
    sheet.cell(row=1, column=2).value = "Test Pass Rate (%)"
    
    passk = metrics['pass@k']
    test_pass_rate = metrics['test_pass_rate']
    
    sheet.cell(row=2, column=1).value = round(passk * 100, 2)
    sheet.cell(row=2, column=2).value = round(test_pass_rate * 100, 2)
    
    tables_dir = os.path.join(SCRIPT_DIR, 'tables')
    os.makedirs(tables_dir, exist_ok=True)
    output_file = os.path.join(tables_dir, 'result_' + file_suffix + '.xlsx')
    workbook.save(output_file)
    
    print("\nEVALUATION RESULTS")
    print(f"  Pass@k:              {passk * 100:.2f}%")
    print(f"  Test Pass Rate:      {test_pass_rate * 100:.2f}%")
    
    return output_file


def store_all_prompts_xlsx(experiment, dataset, model, topn, temperature, used_prompt_type, available_prompt_types=None, suppress_warnings=False):
    """Calculate and store metrics for all prompt types.
    
    Args:
        available_prompt_types: dict mapping prompt_type to list of problems that have this type
                               If None, uses all standard prompt types
    """
    import warnings
    if suppress_warnings:
        warnings.filterwarnings('ignore')
    
    if available_prompt_types:
        all_prompt_types = [pt for pt in ['prompt', 'prompt1a', 'prompt1c', 'prompt1p', 'prompt2ac', 'prompt2ap', 'prompt2cp'] 
                           if pt in available_prompt_types]
    else:
        all_prompt_types = ['prompt', 'prompt1a', 'prompt1c', 'prompt1p', 'prompt2ac', 'prompt2ap', 'prompt2cp']
    
    if used_prompt_type and used_prompt_type not in all_prompt_types:
        all_prompt_types.insert(0, used_prompt_type)
    
    results = {}
    all_metrics = {}
    
    for pt in all_prompt_types:
        try:
            metrics = calculate_metrics(pt)
            results[pt if pt else 'default'] = {
                'pass@k': metrics['pass@k'],
                'test_pass_rate': metrics['test_pass_rate']
            }
            all_metrics[pt] = metrics
        except Exception as e:
            print(f"Warning: Failed to calculate metrics for prompt type '{pt}': {e}")
            results[pt if pt else 'default'] = {
                'pass@k': None,
                'test_pass_rate': None
            }
    
    all_non_prompt_types = ['prompt1a', 'prompt1c', 'prompt1p', 'prompt2ac', 'prompt2ap', 'prompt2cp']
    non_prompt_types = [pt for pt in all_non_prompt_types if pt in all_prompt_types]
    total_passed_cases_all = 0
    total_cases_all = 0
    total_passed_problems_all = 0
    total_problems_all = 0
    
    for pt in non_prompt_types:
        if pt in all_metrics:
            total_passed_cases_all += all_metrics[pt]['total_passed_cases']
            total_cases_all += all_metrics[pt]['total_cases']
            total_passed_problems_all += all_metrics[pt]['num_passed_problems']
            total_problems_all += all_metrics[pt]['num_problems']
    
    if total_cases_all > 0:
        aggregate_test_pass_rate = total_passed_cases_all / total_cases_all
    else:
        aggregate_test_pass_rate = 0.0
    
    if total_problems_all > 0:
        aggregate_passk = total_passed_problems_all / total_problems_all
    else:
        aggregate_passk = 0.0
    
    results['all_non_prompt'] = {
        'pass@k': aggregate_passk,
        'test_pass_rate': aggregate_test_pass_rate
    }
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    sheet.cell(row=1, column=1).value = "Prompt Type"
    sheet.cell(row=1, column=2).value = "Pass@k (%)"
    sheet.cell(row=1, column=3).value = "Test Pass Rate (%)"
    
    row = 2
    for pt_name, metrics in results.items():
        sheet.cell(row=row, column=1).value = pt_name
        if metrics['pass@k'] is not None:
            sheet.cell(row=row, column=2).value = round(metrics['pass@k'] * 100, 2)
            sheet.cell(row=row, column=3).value = round(metrics['test_pass_rate'] * 100, 2)
        else:
            sheet.cell(row=row, column=2).value = "N/A"
            sheet.cell(row=row, column=3).value = "N/A"
        row += 1
    
    tables_dir = os.path.join(SCRIPT_DIR, 'tables')
    os.makedirs(tables_dir, exist_ok=True)
    output_file = os.path.join(tables_dir, f'result_{experiment}_dataset_{dataset}_model_{model}_topn_{topn}_temperature_{temperature}{used_prompt_type}_all_prompts.xlsx')
    workbook.save(output_file)
    
    print("\n" + "="*60)
    print("EVALUATION RESULTS (All Available Prompt Types)")
    print("="*60)
    print(f"Evaluated prompt types: {', '.join(all_prompt_types)}")
    print("="*60)
    for pt_name, metrics in results.items():
        if metrics['pass@k'] is not None:
            print(f"{pt_name:15s}  Pass@k: {metrics['pass@k'] * 100:6.2f}%  Test Pass Rate: {metrics['test_pass_rate'] * 100:6.2f}%")
        else:
            print(f"{pt_name:15s}  Pass@k: N/A      Test Pass Rate: N/A")
    print("="*60)
    
    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-m", "--model", type=str, required=True)
    parser.add_argument("-n", "--topn", type=int, required=True)
    parser.add_argument("-e", "--experiment", type=str, required=True)
    parser.add_argument("-t", "--temperature", type=str, required=True)
    parser.add_argument("-pt", "--prompt_type", type=str, 
                        choices=['', 'prompt1a', 'prompt1c', 'prompt1p', 'prompt2ac', 'prompt2ap', 'prompt2cp'],
                        default='')
    args = parser.parse_args()

    experiment = args.experiment
    dataset = 'HumanEvalComm'
    temperature = args.temperature
    model = args.model
    topn = args.topn
    prompt_type = args.prompt_type

    metrics = calculate_metrics(prompt_type)
    output_file = '%s_dataset_%s_model_%s_topn_%s_temperature_%s%s' % (experiment, dataset, model, topn, temperature, prompt_type)
    store_results_xlsx(metrics, output_file)

